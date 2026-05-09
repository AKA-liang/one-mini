"""
Excel export — record each task's full data to logs/ directory.
"""
from __future__ import annotations

import os
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LOG_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "logs")
os.makedirs(LOG_DIR, exist_ok=True)


def _style_header(ws, row: int, cols: int, fill_color: str = "4472C4"):
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def _style_data(ws, start_row: int, end_row: int, cols: int):
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _auto_width(ws, min_w: int = 8, max_w: int = 40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                lengths.append(len(str(cell.value)))
        width = min(max(max(lengths) + 2 if lengths else min_w, min_w), max_w)
        ws.column_dimensions[col_letter].width = width


def export_task_data(
    task_id: str,
    keywords: str,
    budget: str | None,
    category: str | None,
    chanmama_data: list[dict] | None = None,
    buyin_data: list[dict] | None = None,
    llm_products: list[dict] | None = None,
    finance_data: list[dict] | None = None,
    agent: str = "product_picker",
) -> str:
    """Generate Excel report and return file path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    short_id = task_id[:8] if task_id else "notask"
    filename = f"{agent}_{short_id}_{ts}.xlsx"
    filepath = os.path.join(LOG_DIR, filename)

    wb = openpyxl.Workbook()

    # ═══ Sheet 1: Task Info ═══
    ws1 = wb.active
    ws1.title = "任务信息"
    info = [["字段", "值"], ["任务ID", task_id], ["关键词", keywords],
            ["预算", budget or "无"], ["品类", category or "无"],
            ["时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Agent", agent]]
    for i, row in enumerate(info, 1):
        for j, val in enumerate(row, 1):
            ws1.cell(row=i, column=j, value=val)
    _style_header(ws1, 1, 2)
    _style_data(ws1, 2, len(info), 2)
    _auto_width(ws1)

    # ═══ Sheet 2: Chanmama SPUrank ═══
    if chanmama_data:
        ws2 = wb.create_sheet("蝉妈妈热销SPU")
        headers = ["title", "brand", "sales_volume_text", "sales_amount_text",
                   "live_volume_text", "video_volume_text",
                   "sales_volume_index", "sales_amount_index",
                   "live_amount_index", "video_amount_index",
                   "creator_count", "shop_count", "product_count"]
        for j, h in enumerate(headers, 1):
            ws2.cell(row=1, column=j, value=h)
        _style_header(ws2, 1, len(headers))
        for i, item in enumerate(chanmama_data[:50], 2):
            for j, h in enumerate(headers, 1):
                ws2.cell(row=i, column=j, value=item.get(h, ""))
        _style_data(ws2, 2, min(len(chanmama_data) + 1, 51), len(headers))
        _auto_width(ws2)

    # ═══ Sheet 3: Buyin Search ═══
    if buyin_data:
        ws3 = wb.create_sheet("巨量百应选品广场")
        headers = ["product_name", "price", "commission_rate", "sales", "earn"]
        for j, h in enumerate(headers, 1):
            ws3.cell(row=1, column=j, value=h)
        _style_header(ws3, 1, len(headers), "E74C3C")
        for i, item in enumerate(buyin_data[:50], 2):
            for j, h in enumerate(headers, 1):
                ws3.cell(row=i, column=j, value=item.get(h, ""))
        _style_data(ws3, 2, min(len(buyin_data) + 1, 51), len(headers))
        _auto_width(ws3)

    # ═══ Sheet 4: LLM Selection ═══
    if llm_products:
        ws4 = wb.create_sheet("LLM选品结果")
        headers = ["name", "price", "commission_rate", "monthly_sales",
                   "potential_score", "competition_level", "roi_expectation",
                   "risk_notes", "promotion_suggestion"]
        for j, h in enumerate(headers, 1):
            ws4.cell(row=1, column=j, value=h)
        _style_header(ws4, 1, len(headers), "27AE60")
        for i, item in enumerate(llm_products[:30], 2):
            for j, h in enumerate(headers, 1):
                ws4.cell(row=i, column=j, value=item.get(h, ""))
        _style_data(ws4, 2, min(len(llm_products) + 1, 31), len(headers))
        _auto_width(ws4)

    # ═══ Sheet 5: Finance Analysis ═══
    if finance_data:
        ws5 = wb.create_sheet("财务分析")
        headers = ["name", "selling_price", "commission_rate", "commission_income",
                   "logistics_cost", "ad_cost_per_order", "net_profit_per_order",
                   "profit_margin", "roi", "return_rate", "recommendation", "risk_notes"]
        for j, h in enumerate(headers, 1):
            ws5.cell(row=1, column=j, value=h)
        _style_header(ws5, 1, len(headers), "8E44AD")
        for i, item in enumerate(finance_data[:30], 2):
            for j, h in enumerate(headers, 1):
                ws5.cell(row=i, column=j, value=item.get(h, ""))
            # Color-code recommendations
            rec_cell = ws5.cell(row=i, column=11)
            rec = str(item.get("recommendation", ""))
            if "强烈推荐" in rec:
                rec_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif "推荐" in rec:
                rec_cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            elif "不推荐" in rec:
                rec_cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
            elif "观望" in rec:
                rec_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        _style_data(ws5, 2, min(len(finance_data) + 1, 31), len(headers))
        _auto_width(ws5)

    wb.save(filepath)
    return filepath
